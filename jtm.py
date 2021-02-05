# -*- coding: utf-8 -*-
"""
Created on Sun Jan 17 22:35:38 2021

@author: enrico giampieri
@email: enrico.giampieri@unibo.it
"""

# %% do imports
import os
import contextlib
from sqlite3 import connect, Row
from contextlib import closing
from typing import Any, Mapping, Iterable
import json
from functools import partial
import operator as op
import itertools as it
# I would rather not depend from it, but for now we can't avoid it
import pandas as pd
from openpyxl import load_workbook, Workbook

# %% define the minimum structures

class Table:
    def __init__(self, 
            info: Mapping[str, Any], 
            data: Iterable[Iterable],
            ):
        self.info = info
        self.data = data
        self._columns_name = "columns"
        self._name_name = "name"
        
    @property
    def columns(self) -> Iterable[str]:
        return self.info[self._columns_name]
    
    @property
    def name(self) -> str:
        return self.info[self._name_name]
        
    def __repr__(self) -> str:
        return "{}(info={}, data={})".format(
            self.__class__.__qualname__, 
            self.info, 
            self.data,
            )
    
    def __bool__(self):
        """is truthy if contains values"""
        return len(self.data)!=0
    
    def __eq__(self, other):
        if not isinstance(other, self.__class__):
            return False
        same_info = self.info == other.info
        same_data = self.data == other.data
        return (same_info and same_data)

    def as_pandas(self):
        """convert to a pandas dataframe, the header is used as parameters"""
        df = pd.DataFrame(self.data, columns=self.info['columns'])
        return df
    
    
class DataBase:
    def __init__(self,
            tables: Mapping[str, Table],
            ):
        
        self.tables = tables
        for name, table in self.tables.items():
            assert name == table.name
    
    @property
    def names(self) -> Iterable[str]:
        return list(self.tables.keys())
    
    def __repr__(self) -> str:
        return "{}(tables=[{}])".format(
            self.__class__.__qualname__, 
            self.tables,
            )
    
    def __eq__(self, other):
        if not isinstance(other, self.__class__):
            return False
        if set(self.names) != set(other.names):
            return False
        result = all( self.tables[key]==other.tables[key] for key in self.names )
        return result
    
    def as_pandas(self) -> Iterable:
        tables = [tab.as_pandas() for tab in self.tables]

# %%

def write_into_sql_connection(database, connection):
    for name, table in database.tables.items():
        df = table.as_pandas()
        df.to_sql(name, con=connection, index=False)
        
def write_into_sqlite(database, filename):
    with closing(connect(filename)) as connection:
        write_into_sql_connection(database, connection)

def write_into_excel(database, filename):
    try:
        workbook = load_workbook(filename)
        for name in database.names:
            if name in workbook.sheetnames:
                del workbook[name]
    except FileNotFoundError:    
        # rceate a new one and remove the default worksheet
        workbook = Workbook() 
        del workbook[workbook.sheetnames[0]]
    for name, table in database.tables.items():
        worksheet = workbook.create_sheet(name)
        worksheet.append(table.columns)
        for line in table.data:
            worksheet.append(line)
    workbook.save(filename)

def write_into_jsontable(database, filename):
    with open(filename, "w", encoding="utf8") as outfile:
        for name, table in database.tables.items():
            info = json.dumps(table.info)
            print(info, file=outfile)
            for line in table.data:
                data = json.dumps(line)
                print(data, file=outfile)

def query(database, query, db=":memory:"):
    """return a list of dictionary with the right column names.
    
    It could be converted in a jsontable afterward
    """
    def dict_from_row(row):
        return dict(zip(row.keys(), row))  
    with closing(connect(db)) as connection:
        connection.row_factory = Row
        write_into_sql_connection(database, connection)
        cursor = connection.cursor()
        cursor.execute(query)
        rows = cursor.fetchall()
        result = [dict_from_row(row) for row in rows]
    return result 

# %%
    
def read_from_jsontable(filename):
    # utility functions for type testing later on
    instance_of = lambda types, obj: isinstance(obj, types)
    is_obj_or_arr = partial(instance_of, (dict, list))
    is_arr = partial(instance_of, list)
    # remove the empty lines
    with open(filename, "r", encoding="utf8") as stream:
        lines = filter(None, map(str.strip, stream))
        # transform them in json
        structs = (json.loads(line) for line in lines)
        # keep only objects and arrays
        good_elements = filter(is_obj_or_arr, structs)
        # remove the array that are at the beginning
        good_structs = it.dropwhile(is_arr, good_elements)
        # group together all the objects and all the arrays
        grouped = it.groupby(good_structs, type)
        # drop the data type, keep only the data, we know they are alternating
        grouped_simple = map(list, map(op.itemgetter(1), grouped))
        # pair them, this will also remove any trailing objects with no following arrays
        paired = zip(*([grouped_simple] * 2))
        # if there are multiple objects (technically a mistake) keep only the last one
        paired_single_obj = (Table(info=k[-1], data=d) for k, d in paired)
        # generate the final data structure from the data
        final = {table.info['name']: table for table in paired_single_obj}
    return DataBase(tables=final)

def read_from_excel(filename):
    def _extract_row_data(row):
        return [cell.value for cell in row]
    workbook = load_workbook(filename)
    tables = {}
    for sheetname in workbook.sheetnames:
        worksheet = workbook[sheetname]
        rows = worksheet.iter_rows()
        header = dict(columns=_extract_row_data(next(rows)), name=sheetname)
        data = [_extract_row_data(row) for row in rows]
        header['name'] = sheetname
        result = Table(header, data)
        tables[sheetname] = result
    return DataBase(tables=tables)

def read_from_sqlite(filename):
    def dict_from_row(row):
        return dict(zip(row.keys(), row))
    with closing(connect(filename)) as connection:
        connection.row_factory = Row
        query = "SELECT name FROM sqlite_master WHERE type='table';"
        tablenames = [n[0] for n in connection.execute(query)]
        tables = {}
        for tablename in tablenames:
            cursor = connection.cursor()
            query = "SELECT * FROM {}".format(tablename)
            cursor.execute(query)
            rows = cursor.fetchall()
            rows = [dict_from_row(row) for row in rows]
            columns = list(rows[0].keys())
            info = dict(
                columns=columns, 
                name=tablename,
                )
            data = [[d[c] for c in columns] for d in rows]
            table = Table(info=info, data=data)
            tables[tablename] = table
        db = DataBase(tables)
    return db

# %% useful functions for testing
@contextlib.contextmanager
def _temp_file(filename):
    """context manager to generate temporary files"""
    with contextlib.suppress(FileNotFoundError):
        os.remove(filename)
    yield filename
    with contextlib.suppress(FileNotFoundError):
        os.remove(filename)
        
def _test_data():
    """generate two sample datasets"""
    s1 = Table(
            info={'columns': ['name', 'age'], "name": "ages"}, 
            data=[['alberto', 2], ['barbara', 4], ['carlos', 6]],
            )
    s2 = Table(
            info={'columns': ['name', 'wealth'], "name": "wealths"}, 
            data=[['alberto', 3], ['barbara', 5], ["diana", 7]],
            )
    return s1, s2

def test_roundrobin_jtm_file():
    s1, s2 = _test_data()
    db = DataBase({t.name: t for t in [s1, s2]})
    with _temp_file("mydata.jtm") as jtm_filename:
        write_into_jsontable(db, jtm_filename)
        db2 = read_from_jsontable(jtm_filename)
        assert db == db2
        
def test_roundrobin_xlsx_file():
    s1, s2 = _test_data()
    db = DataBase({t.name: t for t in [s1, s2]})
    with _temp_file("myexcel.xlsx") as xlsx_filename:
        write_into_excel(db, xlsx_filename)
        db2 = read_from_excel(xlsx_filename)
        assert db == db2
        
def test_roundrobin_sqlite_file():
    s1, s2 = _test_data()
    db = DataBase({t.name: t for t in [s1, s2]})
    with _temp_file("mydatabase.db") as sqlite_filename:
        write_into_sqlite(db, sqlite_filename)
        db2 = read_from_sqlite(sqlite_filename)
        assert db == db2
    
def test_sql_query():
    s1, s2 = _test_data()
    db = DataBase({t.name: t for t in [s1, s2]})
    result = query(db, 
        """
        SELECT *
        FROM ages INNER JOIN wealths
        ON ages.name == wealths.name
        """,
        )
    expected = [{'name': 'alberto', 'age': 2, 'wealth': 3}, 
                {'name': 'barbara', 'age': 4, 'wealth': 5},
                ]
    assert result == expected
    
# %%

def main_query(args):
    filename = args.filename
    query = args.query
    db = read_from_jsontable(filename)
    result = query(db, query)
    for line in result:
        print(json.dumps(line))

def main_example(args):
    filename = args.filename
    s1, s2 = _test_data()
    db = DataBase({t.name: t for t in [s1, s2]})
    write_into_jsontable(db, filename)

def main_xlsx2jtm(args):
    source = args.source_filename
    dest = args.destination_filename
    db = read_from_excel(source)
    write_into_jsontable(db, dest)

def main_jtm2xlsx(args):
    source = args.source_filename
    dest = args.destination_filename
    db = read_from_jsontable(source)
    write_into_excel(db, dest)

def main_sqlite2jtm(args):
    source = args.source_filename
    dest = args.destination_filename
    db = read_from_sqlite(source)
    write_into_jsontable(db, dest)

def main_jtm2sqlite(args):
    source = args.source_filename
    dest = args.destination_filename
    db = read_from_jsontable(source)
    write_into_sqlite(db, dest)
    
# %%
if __name__ == '__main__':
    import argparse, sys
    parser = argparse.ArgumentParser()
    parser_subparsers = parser.add_subparsers(
        dest="command",
        title="subcommands",
        description='valid subcommands',
        )
    # querying sql commands
    subparser = parser_subparsers.add_parser(
        'query',
        help="parse a SQL query agains a jtm file",
        )
    if "indentation for query sub command":
        subparser.add_argument(
            "filename",
            help="the file on which to perform the query",
            type=str,
            )
        subparser.add_argument(
            "query",
            help="the file on which to perform the query",
            type=str,
            )
    
    subparser = parser_subparsers.add_parser(
        'example',
        help="generate a simple example jtm files to play with",
        )
    if "indentation for example sub command":
        subparser.add_argument(
            "filename",
            default="example.jtm",
            nargs='?',
            help="the file to fill",
            type=str,
            )

    subparser = parser_subparsers.add_parser(
        'xlsx2jtm',
        help="parse a xlsx file into a jtm",
        )
    if "indentation for sub command":
        subparser.add_argument(
            "source_filename",
            help="",
            type=str,
            )
        subparser.add_argument(
            "destination_filename",
            help="",
            type=str,
            )

    subparser = parser_subparsers.add_parser(
        'jtm2xlsx',
        help="parse a jtm file into a xlsx",
        )
    if "indentation for sub command":
        subparser.add_argument(
            "source_filename",
            help="",
            type=str,
            )
        subparser.add_argument(
            "destination_filename",
            help="",
            type=str,
            )

    subparser = parser_subparsers.add_parser(
        'sqlite2jtm',
        help="parse a sqlite file into a jtm",
        )
    if "indentation for sub command":
        subparser.add_argument(
            "source_filename",
            help="",
            type=str,
            )
        subparser.add_argument(
            "destination_filename",
            help="",
            type=str,
            )

    subparser = parser_subparsers.add_parser(
        'jtm2sqlite',
        help="parse a jtm file into a sqlite",
        )
    if "indentation for sub command":
        subparser.add_argument(
            "source_filename",
            help="",
            type=str,
            )
        subparser.add_argument(
            "destination_filename",
            help="",
            type=str,
            )

    # start the actual parsing and defer
    args = parser.parse_args()  
    if args.command == "query":
        main_query(args)
    elif args.command == "example":
        main_example(args)
    elif args.command == "xlsx2jtm":
        main_xlsx2jtm(args)
    elif args.command == "jtm2xlsx":
        main_jtm2xlsx(args)
    elif args.command == "sqlite2jtm":
        main_sqlite2jtm(args)
    elif args.command == "jtm2sqlite":
        main_jtm2sqlite(args)
    elif args.command is None:
        parser.parse_args(["--help"])
        sys.exit(0)
    else:
        print("wrong command!")
        sys.exit(1)
    
    
# ./jtm.py query test.jtbl "SELECT * from ages INNER JOIN wealths ON ages.name==wealths.name" | vd -f json
