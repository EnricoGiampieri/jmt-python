# -*- coding: utf-8 -*-

# %% import libraries
import json
import gzip
from textwrap import dedent
from io import StringIO, BytesIO

import pandas as pd
import numpy as np

import codecs
import imageio

from sys import getsizeof
import pylab as plt




    
       

# %%

# %% define the functions

class Structure:
    def __init__(self, info=None, data=None):
        self.info = info
        self.data = data if data is not None else []
        self._columns_name = "columns"
        self._name_name = "name"
        
    @property
    def columns(self):
        return self.info[self._columns_name]
    
    @property
    def name(self):
        return self.info[self._name_name]
        
    #
    # object behavior
    #
    
    def __str__(self):
        return "{}(info={}, data={})".format(
            self.__class__.__qualname__, 
            self.info, 
            self.data,
            )
    
    __repr__ = __str__
    
    def __bool__(self):
        """is truthy if contains values"""
        return len(self.data)!=0
    
    def __eq__(self, other):
        if not isinstance(other, self.__class__):
            return False
        same_info = self.info == other.info
        same_data = self.data == other.data
        return (same_info and same_data)
    
    #
    # data format conversion
    #
    
    def as_pandas(self):
        """convert to a pandas dataframe, the header is used as parameters"""
        import pandas as pd
        df = pd.DataFrame(self.data, **self.info)
        return df
    
    def as_jsonlines(self):
        """convert to homogeneous jsonlines, expanded in a series of objects"""
        keys = self.info["columns"]
        result = [dict(zip(keys, values)) for values in self.data]
        return result
    
    #
    # write to file
    #
    
    def to_excel_sheet(self, filename):
        from openpyxl import load_workbook, Workbook
        try:
            workbook = load_workbook(filename)
            if self.name in workbook.sheetnames:
                del workbook[self.name]
        except FileNotFoundError:    
            workbook = Workbook() #
            del workbook[workbook.sheetnames[0]]
        worksheet = workbook.create_sheet(self.name)
        worksheet.append(self.columns)
        for line in self.data:
            worksheet.append(line)
        workbook.save(filename)
        
    def to_sql_table(self, connection):
        df = self.as_pandas()
        df.to_sql(self.info['name'], con=connection)
    
    #
    # read from file
    #
    
    @staticmethod
    def read_from_excel_worksheet(filename, sheetname):
        from openpyxl import load_workbook
        def _extract_row_data(row):
            return [cell.value for cell in row]
        workbook = load_workbook(filename)
        worksheet = workbook[sheetname]
        rows = worksheet.iter_rows()
        header = dict(columns=_extract_row_data(next(rows)), name=sheetname)
        data = [_extract_row_data(row) for row in rows]
        header['name'] = sheetname
        result = Structure(header, data)
        return result
        
    @staticmethod
    def get_tables_from_jsonlines(stream):
        # utility functions for type testing later on
        instance_of = lambda types, obj: isinstance(obj, types)
        is_obj_or_arr = partial(instance_of, (dict, list))
        is_arr = partial(instance_of, list)
        # remove the empty lines
        lines = filter(None, map(str.strip, stream))
        # transform them in json
        structs = (json.loads(line) for line in lines)
        # keep only objects and arrays
        good_elements = filter(is_obj_or_arr, structs)
        # remove the array that are at the beginning
        good_structs = it.dropwhile(is_arr, good_elements)
        # group together all the objects and all the arrays
        grouped = it.groupby(good_structs, type)
        # drop the data type, keep only the data, we know they are alternating
        grouped_simple = map(list, map(op.itemgetter(1), grouped))
        # pair them, this will also remove any trailing objects with no following arrays
        paired = zip(*([grouped_simple] * 2))
        # if there are multiple objects (technically a mistake) keep only the last one
        paired_single_obj = (Structure(info=k[-1], data=d) for k, d in paired)
        # generate the final data structure from the data
        final = {table['info']['name']: table for table in paired_single_obj}
        return final



# %% test di identità
s1a = Structure(
        info={'columns': ['a', 'b']}, 
        data=[[1, 2], [3, 4], [5, 6]],
        )
s1b = Structure(
        info={'columns': ['a', 'b']}, 
        data=[[1, 2], [3, 4], [5, 6]],
        )
s2 = Structure(
        info={'columns': ['a', 'b', 'c']}, 
        data=[[1, 2, 3], [4, 5, 6]],
        )

assert s1a is not s1b
assert s1a == s1b
assert s1a != s2
assert s1b != s2

# %% test per scrivere excel

import os
import contextlib

with contextlib.suppress(FileNotFoundError):
    os.remove("test_1_writing.xlsx")
with contextlib.suppress(FileNotFoundError):
    os.remove("test_2_append.xlsx")
with contextlib.suppress(FileNotFoundError):
    os.remove("test_3_replace.xlsx")

my_data = Structure(
        info={'columns': ['a', 'b'], 'name': 'my data'}, 
        data=[[1, 2], [3, 4], [5, 6]],
        )
my_data.to_excel_sheet("test_1_writing.xlsx")


my_data_2 = Structure(
        info={'columns': ['c', 'd'], 'name': 'my data 2'}, 
        data=[[7, 8], [9, 10], [11, 12]],
        )
my_data.to_excel_sheet("test_2_append.xlsx")
my_data_2.to_excel_sheet("test_2_append.xlsx")

my_data_3 = Structure(
        info={'columns': ['letter', 'number'], 'name': 'my data'}, 
        data=[["a", 1], ["b", 2], ["c", 3]],
        )
my_data.to_excel_sheet("test_3_replace.xlsx")
my_data_3.to_excel_sheet("test_3_replace.xlsx")

# %% reading from excel



def read_from_excel_worksheet(filename, sheetname):
    from openpyxl import load_workbook
    def _extract_row_data(row):
        return [cell.value for cell in row]
    workbook = load_workbook(filename)
    worksheet = workbook[sheetname]
    rows = worksheet.iter_rows()
    header = dict(columns=_extract_row_data(next(rows)), name=sheetname)
    data = [_extract_row_data(row) for row in rows]
    result = Structure(header, data)
    return result

my_data = Structure(
        info={'columns': ['a', 'b'], 'name': 'my data'}, 
        data=[[1, 2], [3, 4], [5, 6]],
        )
my_data.to_excel_sheet("test_1_writing.xlsx")
result = read_from_excel_worksheet("test_1_writing.xlsx", 'my data')
assert result == my_data
# %% primo test
data = """{"columns": ["a", "b"]}
[1, 2]
[3, 4]
[5, 6]
"comment"
{"single object": true}
"""
for structure in reader_header_jsonlines(StringIO(data)):
    print(structure)
print(" ".join(["-*-"]*10))
data = """{"columns": ["a", "b"]}
[1, 2]
[3, 4]
[5, 6]
"""
for structure in reader_header_jsonlines(StringIO(data)):
    print(structure)
    print(structure.as_pandas())
    print(structure.as_jsonlines())

# %% simple tests

#def compress_jsonlines(jsonline):
#    """take a series of dicts and compress them as a jsontable"""
jsonlines = [{'a': 1, 'b': 2}, {'a': 3, 'b': 4}, {'a': 5, 'b': 6}]
keys = [list(line.keys()) for line in jsonlines]
first = set(keys[0])
same_keys = all(set(other)==first for other in keys[1:])
assert same_keys
header = {"columns": list(first)}
data = [[d[k] for k in first] for d in jsonlines]

result = Structure(header, data)
print(result)

# %% second simple test


import itertools as it

def _key_func(structure): 
    """given a dictionary return a sorted tuples of all the keys
    """
    return tuple(sorted(list(structure.keys())))

def jsonlines_objects_to_jsontables(jsonlines):
    for keys, group in it.groupby(jsonlines, _key_func): 
        header = {"columns": list(keys)}
        data = [[d[k] for k in keys] for d in group]
        result = Structure(header, data)
        yield result

jsonlines = [
    {'a': 1, 'b': 2}, 
    {'b': 4, 'a': 3}, 
    {'a': 5, 'b': 6},
    {'a': 1, 'b': 2, 'c': 3},
    {'c': 6, 'a': 4, 'b': 5},
    ]

results = list(jsonlines_objects_to_jsontables(jsonlines))
expected = [
    Structure(
        info={'columns': ['a', 'b']}, 
        data=[[1, 2], [3, 4], [5, 6]],
        ),
    Structure(
        info={'columns': ['a', 'b', 'c']}, 
        data=[[1, 2, 3], [4, 5, 6]],
        ),
    ]

# %% read simply as structure

import itertools as it
import operator as op
from functools import partial
def get_tables(stream):
    # utility functions for type testing later on
    instance_of = lambda types, obj: isinstance(obj, types)
    is_obj_or_arr = partial(instance_of, (dict, list))
    is_arr = partial(instance_of, list)
    # remove the empty lines
    lines = filter(None, map(str.strip, stream))
    # transform them in json
    structs = (json.loads(line) for line in lines)
    # keep only objects and arrays
    good_elements = filter(is_obj_or_arr, structs)
    # remove the array that are at the beginning
    good_structs = it.dropwhile(is_arr, good_elements)
    # group together all the objects and all the arrays
    grouped = it.groupby(good_structs, type)
    # drop the data type, keep only the data, we know they are alternating
    grouped_simple = map(list, map(op.itemgetter(1), grouped))
    # pair them, this will also remove any trailing objects with no following arrays
    paired = zip(*([grouped_simple] * 2))
    # if there are multiple objects (technically a mistake) keep only the last one
    paired_single_obj = ({'info': k[-1], 'data':d} for k, d in paired)
    # generate the final data structure from the data
    final = {table['info']['name']: table for table in paired_single_obj}
    return final

data = StringIO("""
    [1, 2]
    {"columns": ["a", "b"], "name": "foo"}
    [1, {"a": 2}]
    [3, {"a": 4}]
    [5, {"a": 6}]
    {"extraneous object": "remove!"}
    {"columns": ["c", "d"], "name": "bar"}
    [2, [1, 0]]
    [4, [3, 2]]
    [7, [6, 5]]
    {"a": "b"}
    """)
from pprint import pprint
pprint(get_tables(data))


# %%
data = StringIO("""
    [1, 2]
    {"columns": ["a", "b"], "name": "foo"}
    [1, 2]
    [3, 3]
    [5, 4]
    {"extraneous object": "remove!"}
    {"columns": ["c", "d"], "name": "bar"}
    [2, "[1, 0]"]
    [4, "[3, 2]"]
    [7, "[6, 5]"]
    {"a": "b"}
    """)
tables = list(get_tables(data).values())


import os
import contextlib
import sqlite3
with contextlib.suppress(FileNotFoundError):
    os.remove("mydatabase.db")
con = sqlite3.connect('mydatabase.db')
for table in tables:
    df = pd.DataFrame(table['data'], columns=table['info']['columns'])
    df.to_sql(table['info']['name'], con=con)
con.close()
# %%

# %%

# %% NUMPY
"""

[dtype, shape, range, compressed data]

"""
N = 512

data = np.random.randint(2**8, size=(N, N))
# break apart the data
data_bytes = data.tobytes()
dtype = str(data.dtype)
shape = data.shape

# compress and decrompress to check its reversible
compressed = gzip.compress(data_bytes)
data_rebuild = gzip.decompress(compressed)

# rebuild the original array and check it's still equal
reconstructed = np.frombuffer(data_rebuild, dtype=dtype).reshape(*shape)
assert np.allclose(data, reconstructed)
assert bytes(list(compressed)) == compressed
# %%
# check if the data transposed is significantly different in size
data_bytes = data.tobytes()
compressed = gzip.compress(data_bytes)
print(len(data_bytes)/len(compressed))
data_bytes_T = data.T.tobytes()
compressed_T = gzip.compress(data_bytes_T)
print(len(data_bytes_T)/len(compressed_T))

# %%
# compress the array as a whole or break it apart, how much does it change size?
b0 = gzip.compress(data.tobytes())

b1 = gzip.compress(data[:N//2, :N//2].tobytes())
b2 = gzip.compress(data[:N//2, N//2:].tobytes())
b3 = gzip.compress(data[N//2:, :N//2].tobytes())
b4 = gzip.compress(data[N//2:, N//2:].tobytes())

L1 = len(b0)
L2 = sum(map(len, [b1, b2, b3, b4]))

print(L2/L1)
# %%
string_representation = repr(list(compressed)).replace(' ', '').encode('utf8')
print(len(string_representation)/len(compressed))
print(len(string_representation)/len(data_bytes))

b = json.loads(string_representation)
b = bytes(b)
assert compressed == b
# %%

base_64_encoded = codecs.encode(compressed, 'base64')
base_64_encode_simplified = base_64_encoded.decode('utf8').replace('\n', '').encode('utf8')
rebuilt = codecs.decode(base_64_encoded, 'base64')
assert rebuilt==compressed
# %%
len(base_64_encode_simplified)/len(data_bytes)
len(compressed)/len(data_bytes)

# %%
"""there shdoul be a pair of functions, both getting bytes in and out
in the expanded and compressed form, no other parameters!
"""

def gzip_compress(arr):
    data_bytes = arr.tobytes()
    return gzip.compress(data_bytes)

def gzip_decompress(data_bytes):
    return gzip.decompress(data_bytes)


def img_compress(format='jpeg', **kwargs):
    def compress(arr):
        output = BytesIO()
        imageio.imsave(output, arr, format=format, **kwargs)
        output = output.getvalue()
        return output
    return compress

def img_decompress(format='jpeg', **kwargs):
    def decompress(data_bytes):
        re_read = imageio.imread(BytesIO(data_bytes), format=format, **kwargs)
        return re_read
    return decompress
# %%
# def _list_to_tuple(seq):
#     if isinstance(seq, list):
#         return tuple([_list_to_tuple(s) for s in seq])
#     else:
#         return seq

# def loads_numpy_dtype_from_json(str_data):
#     #return np.dtype(list(_list_to_tuple(json.loads(str_data))))
#     return np.dtype(list(_list_to_tuple(str_data)))

# %%

def compress_array(arr, compression=gzip.compress):
    dtype = str(arr.dtype)
    shape = arr.shape
    compressed = compression(arr)
    base_64_encoded = codecs.encode(compressed, 'base64')
    encoded = (base_64_encoded.decode('utf8')
                .replace('\n', ''))
    return dtype, shape, encoded

def decompress_array(dtype, shape, base64_str, decompression=gzip.decompress):
    base64_bytes = base64_str.encode('utf8')
    compressed = codecs.decode(base64_bytes, 'base64')
    data_rebuild = decompression(compressed)
    extracted = np.frombuffer(data_rebuild, dtype=dtype).reshape(*shape)
    return extracted

# %%
N = 2**10
data = np.random.randint(2**8, size=(N, N), dtype='uint8')
data = np.arange(N, dtype='float')#.reshape(N, N)

dtype, shape, base64_str = compress_array(data)
rebuilt = decompress_array(dtype, shape, base64_str)
assert np.allclose(data, rebuilt)
print(getsizeof(base64_str)/getsizeof(data.tobytes()))

# %%
filename = "C:/Users/enrico/Documents/avatar_mani.jpg" # 21 kb
image = imageio.imread(filename)

# %%
dtype, shape, base64_str = compress_array(image, img_compress('png'))
rebuilt = decompress_array(dtype, shape, base64_str, img_decompress('png'))
assert np.allclose(image, rebuilt)
print(getsizeof(base64_str)/getsizeof(image.tobytes()))

# %%
params = dict(format='jpeg', compress=True)
dtype, shape, base64_str = compress_array(image, img_compress(**params))
rebuilt = decompress_array(dtype, shape, base64_str, img_decompress('jpeg'))
#assert np.allclose(image, rebuilt)
print(getsizeof(base64_str)/getsizeof(image.tobytes()))

# %%
# class DecoderError(TypeError):
#     pass

# def reader_header_jsonlines_old(stream, mapper):
#     last_header = None
#     for line in stream:
#         struct = json.loads(line.strip())
#         if isinstance(struct, dict):
#             last_header = struct
#         elif isinstance(struct, list):
#             yield mapper(struct, last_header)
#         else:
#             raise DecoderError("object or list extected")
            
# def column_mapper(struct, header):
#     default_header = list(map(str, range(len(struct))))
#     columns = header.get("columns", default_header)
#     struct_dict = dict(zip(columns, struct))
#     other_columns = {k:v for k,v in header.items() if k!="columns"}
#     struct_dict.update(other_columns)
#     return struct_dict


# %% generate data
data = np.random.randint(1, 6, size=4, dtype='uint8')
dtype, shape, base64_str = compress_array(data)

d0 = {"metadata": True, "coding": {"0": "Male", "1": "Female"}}
d1 = {"columns": ["dtype", "shape", "data"], "data_compression": "gzip"}
d2 = ["int8", [2, 2], "H4sIAOATk14C/2NgZGIGABOGuYsEAAAA"]
d3 = [dtype, shape, base64_str]
d4 = {"data_compression": "gzip"}
d5 = ["int8", [2, 2], "H4sIAOATk14C/2NgZGIGABOGuYsEAAAA"]
d6 = {"metadata": False}
d =  [d0, d1, d2, d3, d4, d5, d6]
jsons = list(map(json.dumps, d))
s = "\n".join(jsons)
positions = np.r_[0, np.cumsum(list(map(len, jsons)))]
# %% test the basic data extraction
o = StringIO(s)
structs = list(reader_header_jsonlines(o))
assert len(structs)==4
assert structs[0].info==d0
assert structs[1].info==d1
assert structs[2].info==d4
# %% te
o = StringIO(s)
for line in o:
    end_pos = o.tell()
    start_pos = end_pos - len(line)
    print(start_pos, end_pos)
    #print(line)


# %%
import os
 
def read_reverse_order(read_obj, encoding='utf8'):
    # adepted from
    # https://thispointer.com/python-read-a-file-in-reverse-order-line-by-line/
    # Move the cursor to the end of the file
    read_obj.seek(0, os.SEEK_END)
    # Create a buffer to keep the last read line
    buffer = bytearray()
    # Get the current position of pointer i.e eof
    last_pointer_location = read_obj.tell()
    # Loop till pointer reaches the top of the file
    for pointer_location in range(last_pointer_location, -1, -1):
        # Move the file pointer to the location pointed by pointer_location
        read_obj.seek(pointer_location)
        # read that byte / character
        new_byte = read_obj.read(1)
        # If the read byte is new line character then it means one line is read
        if new_byte == b'\n':
            # Fetch the line from buffer and yield it
            yield buffer.decode(encoding)[::-1]
            # Reinitialize the byte array to save next line
            buffer = bytearray()
        else:
            # If last read character is not eol then add it in buffer
            buffer.extend(new_byte)
    # As file is read completely, if there is still data in buffer, then its the first line.
    if len(buffer) > 0:
        # Yield the first line too
        yield buffer.decode(encoding)[::-1]

read_file = BytesIO(s.encode('utf8'))
for line in read_reverse_order(read_file):
    print(line)


# %%


